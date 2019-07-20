# -*-coding:utf-8 -*-
import openpyxl as opx
import os.path
from copy import copy
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment


class Grade():
    def __init__(self, subject):
        self.subject = subject
        self.selection = 0
        self.score = []
        self.div_low = []
        self.div_up = []

    def sort(self):
        self.score.sort(reverse=True)

    def divide(self, bounds):
        self.sort()
        for i in range(0, 8):
            self.div_low.append(
                self.score[int(self.selection * bounds[i]) - 1])  # 下标从0开始因此减去1
            if i == 0:
                self.div_up.append(self.score[0])
            else:
                up = self.score[int(self.selection * bounds[i - 1])]
                for i in self.score:
                    if i < up:
                        up = i
                        break
                self.div_up.append(up)


def statistic(worksheet, grades, bounds):
    """统计各科的选择人数、每个人的选科组合，剔除非法数据(例如某考生仅考了两科)，并计算各分数区间。"""
    worksheet.insert_cols(14)
    worksheet.cell(2, 14).value = "选课"
    row = 3
    illegal = 0
    while row <= worksheet.max_row:
        selection = ""
        for col in range(8, 14):
            if worksheet.cell(row, col).value:
                if worksheet.cell(2, col).value == "历史":
                    selection += "史"
                else:
                    selection += (worksheet.cell(2, col).value)[0:1]
        if len(selection) == 3:
            worksheet.cell(row, 14).value = selection
            row += 1
        else:
            worksheet.delete_rows(row)
            illegal += 1
    print("剔除 " + str(illegal) + " 条非法数据后：")
    for col in range(8, 14):
        subject = worksheet.cell(2, col).value
        for row in range(3, worksheet.max_row + 1):
            if worksheet.cell(row, col).value:
                grades[subject].selection += 1
                grades[subject].score.append(
                    float(worksheet.cell(row, col).value))
        print("共有 " + str(grades[subject].selection) + " 人选择" + subject + ",")
        grades[subject].divide(bounds)
    print("共有 " + str(worksheet.max_row - 2) + " 人成绩有效.")


def export_div(grades):
    divbook = opx.load_workbook(r"分数区间.xlsx")
    divsheet = divbook.active
    for row in range(3, 9):
        subject = divsheet.cell(row, 2).value
        cnt = 0
        for col in range(3, divsheet.max_column + 1, 2):
            divsheet.cell(row, col).value = grades[subject].div_up[cnt]
            cnt += 1
        cnt = 0
        for col in range(4, divsheet.max_column + 1, 2):
            divsheet.cell(row, col).value = grades[subject].div_low[cnt]
            cnt += 1
    divbook.save(r"分数区间.xlsx")


def calc(grade, standard, origin, division):
    if grade.div_up[division] == origin:
        return standard[division][1]
    if grade.div_low[division] == origin:
        return standard[division][0]
    temp = float(grade.div_up[division] - origin) / \
        float(origin - grade.div_low[division])
    trans = float(standard[division][1] +
                  standard[division][0] * temp) / float(temp + 1)
    return trans


def export_trans(transbook, grades, standard):
    transheet = transbook.active
    col = 9
    while col <= 19:
        transheet.insert_cols(col)
        transheet.cell(2, col).value = transheet.cell(
            2, col - 1).value + "(转换)"
        col += 2
    for row in range(3, transheet.max_row):
        for col in range(9, 20, 2):
            if (transheet.cell(row, col - 1).value):
                subject = transheet.cell(2, col - 1).value
                origin = float(transheet.cell(row, col - 1).value)
                division = 0
                for i in range(0, 8):
                    up = grades[subject].div_up[i]
                    low = grades[subject].div_low[i]
                    if origin >= low and origin <= up:
                        division = i
                        break
                transheet.cell(row, col).value = calc(
                    grades[subject], standard, origin, division)
    transheet.cell(2, 21).value = "原始总分"
    transheet.cell(2, 22).value = "转换总分"
    transheet.cell(2, 23).value = "排名"
    formula(transheet)
    format(transheet)
    transbook.save(r"转换成绩.xlsx")


def format(worksheet):
    font = Font("Arial", 10)
    border = Border(left=Side(border_style="thin",
                              color='FF000000'),
                    right=Side(border_style="thin",
                               color='FF000000'),
                    top=Side(border_style="thin",
                             color='FF000000'),
                    bottom=Side(border_style="thin",
                                color='FF000000'),
                    diagonal=Side(border_style="thin",
                                  color='FF000000'),
                    diagonal_direction=0,
                    outline=Side(border_style="thin",
                                 color='FF000000'),
                    vertical=Side(border_style="thin",
                                  color='FF000000'),
                    horizontal=Side(border_style="thin",
                                    color='FF000000'))
    align = Alignment(horizontal="center", vertical="center")
    head_fill = PatternFill("solid", "ffbfbfbf")
    worksheet.merge_cells("A1:W1")
    for row in worksheet.rows:
        for cell in row:
            if cell.row <= 2:
                cell.fill = head_fill
            cell.font = font
            cell.border = border
            cell.alignment = align


def formula(worksheet):
    for row in range(3, worksheet.max_row):
        worksheet.cell(
            row, 21).value = "=SUM(E{0}:G{0},H{0},J{0},L{0},N{0},P{0},R{0})".format(row)
        worksheet.cell(
            row, 22).value = "=SUM(E{0}:G{0},I{0},K{0},M{0},O{0},Q{0},S{0})".format(row)
        worksheet.cell(row, 23).value = "=ROW()-2"


def run():
    print("正在验证文件完整性...")
    if os.path.exists(r"原始成绩.xlsx") == False or os.path.exists(r"分数区间.xlsx") == False:
        input("缺少必要文件, 请重新下载!\n按任意键退出:")
        exit()
    input("验证成功,\n请将原始成绩粘贴到运行目录下的\"原始成绩.xlsx\"表格中,\n按任意键继续：")
    workbook = opx.load_workbook(r"原始成绩.xlsx")
    worksheet = workbook.active
    worksheet.unmerge_cells("A1:N1")
    grades = {'物理': Grade('物理'), '化学': Grade('化学'), '生物': Grade(
        '生物'), '历史': Grade('历史'), '政治': Grade('政治'), '地理': Grade('地理')}
    bounds = (0.03, 0.10, 0.26, 0.50, 0.74, 0.90, 0.97, 1.00)
    standard = ((91, 100), (81, 90), (71, 80), (61, 70),
                (51, 60), (41, 50), (31, 40), (21, 30))
    print("正在统计...")
    statistic(worksheet, grades, bounds)
    workbook.save(r"合法原始.xlsx")
    print("正在计算赋分区间...")
    export_div(grades)
    print("正在赋分，请等待...")
    export_trans(workbook, grades, standard)
    print("导出转换成绩成功...\n请到运行目录下查看\"转换成绩.xlsx\".")


run()
